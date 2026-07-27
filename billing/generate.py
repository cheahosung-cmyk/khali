#!/usr/bin/env python3
"""웰스타임 관리비 생성 스크립트.

사용법:
    python3 billing/generate.py 2026-07

읽는 파일:
    billing/building.json          세대 목록·단가·계좌 (거의 안 바뀜)
    billing/months/2026-07.json    이번 달 입력값 (고지서에서 뽑은 숫자 + 검침값)
    billing/ledger.json            세대별 미납 잔액 (전월 것을 자동으로 가져옴)

만드는 것:
    billing/out/2026-07/관리비_2026-07.xlsx   계산표 + 미납관리 + 세대별 고지서 시트
    billing/out/2026-07/고지서_2026-07.md     세대별 고지서 문구 (문자 발송용)
    표준출력                                   세대별 청구액 표 + 검산 결과
    billing/ledger.json                        이번 달 말 미납 잔액 갱신 (재실행해도 안전)
"""

import json
import sys
from pathlib import Path

BASE = Path(__file__).resolve().parent


def krw(n):
    return f"{n:,}원"


def prev_month(ym):
    y, m = map(int, ym.split("-"))
    y, m = (y - 1, 12) if m == 1 else (y, m - 1)
    return f"{y:04d}-{m:02d}"


def load(month):
    building = json.loads((BASE / "building.json").read_text(encoding="utf-8"))
    month_file = BASE / "months" / f"{month}.json"
    if not month_file.exists():
        sys.exit(f"입력 파일이 없습니다: {month_file}")
    inputs = json.loads(month_file.read_text(encoding="utf-8"))
    ledger_file = BASE / "ledger.json"
    ledger = json.loads(ledger_file.read_text(encoding="utf-8")) if ledger_file.exists() else {}
    return building, inputs, ledger


def calculate(building, inputs, ledger, month):
    단가 = building["단가"]
    연체율 = 단가["연체료율"]
    수도 = inputs["수도"]
    전기 = inputs["전기"]
    사용량 = inputs.get("세대별_수도사용량_t", {})
    납부 = inputs.get("이번달_납부액", {})
    전월미납 = ledger.get(prev_month(month), {})

    # '퇴거': 'YYYY-MM'이 있는 세대는 그 달부터 부과에서 빠진다 (문자열 비교로 충분)
    재실 = [s for s in building["세대"] if not s.get("퇴거") or month < s["퇴거"]]
    지분평수합 = sum(s.get("지분평수", s["분양평수"]) for s in 재실)
    수도단가 = 수도["납기요금"] / 수도["전체사용량_t"] if 수도["전체사용량_t"] else 0
    전기총액 = 전기["기본요금"] + 전기["전기요금"] + 전기["전력기금"]

    rows = []
    for s in 재실:
        이름, 평수 = s["이름"], s["분양평수"]
        지분 = s.get("지분평수", 평수) / 지분평수합
        세대수도 = round(수도단가 * 사용량.get(이름, 0))
        rows.append({
            "이름": 이름, "호": s.get("호", ""), "평수": 평수, "지분": 지분,
            "청소비품": s.get("청소비품", 0), "부가세적용": s.get("부가세적용", True),
            "수도사용량": 사용량.get(이름, 0), "세대상수도료": 세대수도,
        })

    공용수도 = 수도["납기요금"] - sum(r["세대상수도료"] for r in rows)

    for r in rows:
        r["공동수도료"] = round(공용수도 * r["지분"])
        r["전기료"] = round(전기총액 * r["지분"])
        r["일반관리비"] = round(r["평수"] * 단가["평당관리비"])
        r["부가세"] = round(r["일반관리비"] * 단가["부가세율"]) if r["부가세적용"] else 0
        r["당월부과액"] = (r["세대상수도료"] + r["공동수도료"] + r["전기료"]
                          + r["일반관리비"] + r["부가세"] + r["청소비품"])
        r["미납액"] = 전월미납.get(r["이름"], 0)
        r["미납연체료"] = int(r["미납액"] * 연체율)  # 실제 청구서와 동일하게 원 미만 절사
        r["납기내금액"] = r["당월부과액"] + r["미납액"] + r["미납연체료"]
        r["납기후금액"] = r["납기내금액"] + int(r["당월부과액"] * 연체율)
        r["납부액"] = 납부.get(r["이름"], 0)
        r["기말미납"] = r["미납액"] + r["당월부과액"] - r["납부액"]

    # 퇴거한 세대는 부과 없이 미수금 잔액만 이월한다
    퇴거 = []
    for s in building["세대"]:
        if s.get("퇴거") and month >= s["퇴거"]:
            잔액 = 전월미납.get(s["이름"], 0)
            낸돈 = 납부.get(s["이름"], 0)
            if 잔액 or 낸돈:
                퇴거.append({"이름": s["이름"], "미납액": 잔액, "납부액": 낸돈,
                             "기말미납": 잔액 - 낸돈,
                             "비고": f"퇴거({s['퇴거']}) 미수금 추적, 부과 없음"})

    meta = {"지분평수합": 지분평수합, "수도단가": 수도단가, "공용수도": 공용수도,
            "전기총액": 전기총액, "전월미납": 전월미납}
    return rows, 퇴거, meta


def verify(rows, inputs, meta):
    """합계 검산: 수도 배분 합이 납기요금과 일치하는지, 전기 반올림 오차는 몇 원인지."""
    수도합 = sum(r["세대상수도료"] + r["공동수도료"] for r in rows)
    전기합 = sum(r["전기료"] for r in rows)
    lines = []
    수도오차 = 수도합 - inputs["수도"]["납기요금"]
    전기오차 = 전기합 - meta["전기총액"]
    lines.append(f"- 수도 배분합 {krw(수도합)} vs 납기요금 {krw(inputs['수도']['납기요금'])}"
                 f" (오차 {수도오차:+,}원, 세대별 반올림 때문)")
    lines.append(f"- 전기 배분합 {krw(전기합)} vs 전기 총액 {krw(meta['전기총액'])}"
                 f" (오차 {전기오차:+,}원)")
    ok = abs(수도오차) <= len(rows) and abs(전기오차) <= len(rows)
    return lines, ok


def print_summary(rows, 퇴거, meta, inputs, month):
    y, m = month.split("-")
    print(f"# {int(y)}년 {int(m)}월 웰스타임 관리비\n")
    print("| 세대 | 일반관리비 | 부가세 | 청소비품 | 세대상수도 | 공동수도 | 전기료 | 당월부과액 | 미납액 | 연체료 | 납기내금액 | 납기후금액 |")
    print("|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|")
    for r in rows:
        print(f"| {r['이름']} | {r['일반관리비']:,} | {r['부가세']:,} | {r['청소비품']:,} "
              f"| {r['세대상수도료']:,} | {r['공동수도료']:,} | {r['전기료']:,} "
              f"| {r['당월부과액']:,} | {r['미납액']:,} | {r['미납연체료']:,} "
              f"| **{r['납기내금액']:,}** | {r['납기후금액']:,} |")
    tot = {k: sum(r[k] for r in rows) for k in
           ["일반관리비", "부가세", "청소비품", "세대상수도료", "공동수도료", "전기료",
            "당월부과액", "미납액", "미납연체료", "납기내금액", "납기후금액"]}
    print(f"| 합계 | {tot['일반관리비']:,} | {tot['부가세']:,} | {tot['청소비품']:,} "
          f"| {tot['세대상수도료']:,} | {tot['공동수도료']:,} | {tot['전기료']:,} "
          f"| {tot['당월부과액']:,} | {tot['미납액']:,} | {tot['미납연체료']:,} "
          f"| **{tot['납기내금액']:,}** | {tot['납기후금액']:,} |")
    for t in 퇴거:
        print(f"\n※ {t['이름']}: 미수금 {krw(t['기말미납'])} ({t['비고']})")
    print("\n검산:")
    for line in verify(rows, inputs, meta)[0]:
        print(line)


def make_xlsx(rows, 퇴거, meta, inputs, building, month, outdir):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    head_fill = PatternFill("solid", fgColor="D9E1F2")
    sum_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    num = "#,##0"

    def style_row(ws, r, n_cols, bold=False, fill=None):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if bold:
                cell.font = Font(bold=True)
            if fill:
                cell.fill = fill
            if isinstance(cell.value, (int, float)) and c > 1:
                cell.number_format = num

    y, m = month.split("-")
    title = f"{int(y)}년 {int(m)}월"

    ws = wb.active
    ws.title = "관리비계산"
    ws["A1"] = f"웰스타임 관리비 계산표 — {title}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = (f"수도: 전체 {inputs['수도']['전체사용량_t']}t, 납기요금 {krw(inputs['수도']['납기요금'])}, "
                f"기간 {inputs['수도']['사용기간']} / 전기 총액 {krw(meta['전기총액'])}, "
                f"기간 {inputs['전기']['사용기간']} / 수도 단가 {meta['수도단가']:,.1f}원/t, "
                f"공용수도 {krw(meta['공용수도'])}")
    headers = ["세대명", "분양평수", "공용지분율", "수도사용량(t)", "세대상수도료", "공동수도료",
               "전기료", "일반관리비", "부가세", "청소비품", "당월부과액", "미납액",
               "미납연체료", "납기내금액", "납기후금액"]
    ws.append([])
    ws.append(headers)
    hr = ws.max_row
    style_row(ws, hr, len(headers), bold=True, fill=head_fill)
    keys = ["이름", "평수", "지분", "수도사용량", "세대상수도료", "공동수도료", "전기료",
            "일반관리비", "부가세", "청소비품", "당월부과액", "미납액", "미납연체료",
            "납기내금액", "납기후금액"]
    for r in rows:
        ws.append([r[k] for k in keys])
        style_row(ws, ws.max_row, len(headers))
        ws.cell(row=ws.max_row, column=3).number_format = "0.0000"
    합 = ["합 계", sum(r["평수"] for r in rows), sum(r["지분"] for r in rows),
          sum(r["수도사용량"] for r in rows)] + \
        [sum(r[k] for r in rows) for k in keys[4:]]
    ws.append(합)
    style_row(ws, ws.max_row, len(headers), bold=True, fill=sum_fill)
    ws.cell(row=ws.max_row, column=3).number_format = "0.0000"
    ws.column_dimensions["A"].width = 14
    for col in "DEFGHIJKLMNO":
        ws.column_dimensions[col].width = 12

    ws2 = wb.create_sheet("미납관리")
    ws2["A1"] = f"미납 관리 — {title}"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A2"] = "연체료는 전월 이월미납액에 1.5% 한 번만 부과(청구서 반영). 원금에는 더하지 않음(복리 아님)."
    ws2.append([])
    ws2.append(["세대명", "전월 이월미납액", "미납연체료(1.5%)", "이번달 당월부과액",
                "이번달 납부액", "이번달말 미납잔액", "비고"])
    style_row(ws2, ws2.max_row, 7, bold=True, fill=head_fill)
    for r in rows:
        ws2.append([r["이름"], r["미납액"], r["미납연체료"], r["당월부과액"],
                    r["납부액"], r["기말미납"], ""])
        style_row(ws2, ws2.max_row, 7)
    for t in 퇴거:
        ws2.append([t["이름"], t["미납액"], 0, 0, t["납부액"], t["기말미납"], t["비고"]])
        style_row(ws2, ws2.max_row, 7)
    전체 = rows + 퇴거
    ws2.append(["합 계", sum(x["미납액"] for x in 전체), "",
                sum(r["당월부과액"] for r in rows), sum(x["납부액"] for x in 전체),
                sum(x["기말미납"] for x in 전체), ""])
    style_row(ws2, ws2.max_row, 7, bold=True, fill=sum_fill)
    ws2.column_dimensions["A"].width = 14
    for col in "BCDEF":
        ws2.column_dimensions[col].width = 15
    ws2.column_dimensions["G"].width = 30

    항목 = [("일반관리비", "일반관리비"), ("관리비 부가세", "부가세"), ("청소비품", "청소비품"),
           ("세대상수도료", "세대상수도료"), ("공동수도료", "공동수도료"), ("전기료", "전기료")]
    for r in rows:
        ws3 = wb.create_sheet(r["이름"][:28])
        ws3["A1"] = f"{title} 관리비 고지서"
        ws3["A1"].font = Font(bold=True, size=14)
        ws3["A2"] = f"세대: {r['이름']} ({r['호']}호, 분양 {r['평수']}평)"
        ws3["A2"].font = Font(bold=True)
        ws3.append([])
        ws3.append(["항목", "금액"])
        style_row(ws3, ws3.max_row, 2, bold=True, fill=head_fill)
        for 라벨, k in 항목:
            ws3.append([라벨, r[k]])
            style_row(ws3, ws3.max_row, 2)
        ws3.append(["당월부과액", r["당월부과액"]])
        style_row(ws3, ws3.max_row, 2, bold=True)
        ws3.append(["미납액", r["미납액"]])
        ws3.append(["미납연체료", r["미납연체료"]])
        style_row(ws3, ws3.max_row - 1, 2)
        style_row(ws3, ws3.max_row, 2)
        ws3.append(["납기내금액", r["납기내금액"]])
        style_row(ws3, ws3.max_row, 2, bold=True, fill=sum_fill)
        ws3.append(["납기후금액", r["납기후금액"]])
        style_row(ws3, ws3.max_row, 2)
        ws3.append([])
        ws3.append(["납부기한", building["납부기한"]])
        ws3.append(["입금계좌", building["입금계좌"]])
        ws3.append(["발행", building["발행처"]])
        ws3.column_dimensions["A"].width = 16
        ws3.column_dimensions["B"].width = 34
        for rr in range(1, ws3.max_row + 1):
            ws3.cell(row=rr, column=2).alignment = Alignment(horizontal="right")

    path = outdir / f"관리비_{month}.xlsx"
    wb.save(path)
    return path


def make_notices(rows, building, inputs, month, outdir):
    y, m = month.split("-")
    title = f"{int(y)}년 {int(m)}월"
    parts = [f"# {title} 세대별 고지서 문구\n"]
    for r in rows:
        미납줄 = ""
        if r["미납액"]:
            미납줄 = (f"- 미납액: {krw(r['미납액'])}\n"
                      f"- 미납연체료(1.5%): {krw(r['미납연체료'])}\n")
        parts.append(f"""## {r['이름']} ({r['호']}호)

[{building['발행처']}] {title} 관리비 안내드립니다.

- 일반관리비: {krw(r['일반관리비'])} (분양 {r['평수']}평 × {building['단가']['평당관리비']:,}원)
- 부가세: {krw(r['부가세'])}
- 청소비품: {krw(r['청소비품'])}
- 세대상수도료: {krw(r['세대상수도료'])} (사용 {r['수도사용량']}t, 기간 {inputs['수도']['사용기간']})
- 공동수도료: {krw(r['공동수도료'])}
- 전기료: {krw(r['전기료'])} (기간 {inputs['전기']['사용기간']})
- 당월부과액: {krw(r['당월부과액'])}
{미납줄}
납기내금액: **{krw(r['납기내금액'])}** (납부기한: {building['납부기한']})
납기후금액: {krw(r['납기후금액'])}

입금계좌: {building['입금계좌']}
""")
    path = outdir / f"고지서_{month}.md"
    path.write_text("\n".join(parts), encoding="utf-8")
    return path


def update_ledger(rows, 퇴거, month):
    ledger_file = BASE / "ledger.json"
    ledger = json.loads(ledger_file.read_text(encoding="utf-8")) if ledger_file.exists() else {}
    ledger[month] = {x["이름"]: x["기말미납"] for x in rows + 퇴거 if x["기말미납"]}
    ledger_file.write_text(json.dumps(ledger, ensure_ascii=False, indent=2) + "\n",
                           encoding="utf-8")


def main():
    if len(sys.argv) != 2:
        sys.exit("사용법: python3 billing/generate.py 2026-07")
    month = sys.argv[1]
    building, inputs, ledger = load(month)
    rows, 퇴거, meta = calculate(building, inputs, ledger, month)

    _, ok = verify(rows, inputs, meta)
    if not ok:
        sys.exit("검산 실패: 배분 합계가 반올림 허용 범위를 벗어났습니다. 입력값을 확인하세요.")

    outdir = BASE / "out" / month
    outdir.mkdir(parents=True, exist_ok=True)
    xlsx = make_xlsx(rows, 퇴거, meta, inputs, building, month, outdir)
    notices = make_notices(rows, building, inputs, month, outdir)
    update_ledger(rows, 퇴거, month)

    print_summary(rows, 퇴거, meta, inputs, month)
    print(f"\n생성 완료:\n- {xlsx}\n- {notices}\n- ledger.json에 {month} 미납잔액 기록")
    if inputs.get("비고"):
        print(f"\n주의: {inputs['비고']}")


if __name__ == "__main__":
    main()
