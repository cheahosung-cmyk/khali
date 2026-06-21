"""엑셀(.xlsx) 고지서 출력 (선택적 기능).

openpyxl 이 설치돼 있을 때만 동작한다. (pip install openpyxl)
설치돼 있지 않으면 안내 메시지와 함께 RuntimeError 를 발생시킨다.
"""

from __future__ import annotations

from typing import Dict, List, Optional

from .models import Settlement


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as e:  # pragma: no cover
        raise RuntimeError(
            "엑셀 출력에는 openpyxl 이 필요합니다. "
            "'pip install openpyxl' 설치 후 다시 실행하세요."
        ) from e
    return openpyxl


def write_bills_xlsx(path: str, settlement: Settlement,
                     prev: Optional[Dict[str, int]] = None) -> None:
    """호실별 고지서를 엑셀(.xlsx)로 저장한다."""
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    item_names = [item.name for item in settlement.items]
    use_vat = settlement.vat_rate > 0
    use_status = settlement.has_vacant
    use_prev = prev is not None

    header = ["호실", "평수"]
    if use_status:
        header.append("상태")
    header += item_names
    header += ["공급가액", "부가세", "청구합계"] if use_vat else ["합계"]
    if use_prev:
        header += ["전월합계", "증감"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "관리비고지서"

    head_fill = PatternFill("solid", fgColor="305496")
    head_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill("solid", fgColor="DDEBF7")
    bold = Font(bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center")

    ws.append(header)
    for c in ws[1]:
        c.fill = head_fill
        c.font = head_font
        c.alignment = center
        c.border = border

    def _amount_cells(bill):
        cells: List = [bill.charges.get(n, 0) for n in item_names]
        if use_vat:
            if bill.unit.occupied:
                cells += [bill.supply, settlement.vat(bill), settlement.billed_total(bill)]
            else:
                cells += [bill.supply, "-", "-"]
        else:
            cells += [bill.supply]
        if use_prev:
            if bill.unit.occupied and bill.unit.name in prev:
                pv = prev[bill.unit.name]
                cells += [pv, settlement.claim(bill) - pv]
            else:
                cells += ["", ""]
        return cells

    def _add(bill, status=None):
        row = [bill.unit.name, bill.unit.area]
        if use_status:
            row.append(status)
        row += _amount_cells(bill)
        ws.append(row)

    for bill in settlement.tenant_bills:
        _add(bill, "임차")
    for bill in settlement.vacant_bills:
        _add(bill, "공실(건물주부담)")

    # 합계 행
    base = ["임차인 합계", None] + ([None] if use_status else [])
    tail = [sum(b.charges.get(n, 0) for b in settlement.tenant_bills) for n in item_names]
    if use_vat:
        tail += [settlement.supply_billed_total, settlement.vat_total, settlement.billed_grand_total]
    else:
        tail += [settlement.supply_billed_total]
    if use_prev:
        prev_sum = sum(prev.get(b.unit.name, 0) for b in settlement.tenant_bills)
        cur_sum = sum(settlement.claim(b) for b in settlement.tenant_bills if b.unit.name in prev)
        tail += [prev_sum, cur_sum - prev_sum]
    total_row_idx = ws.max_row + 1
    ws.append(base + tail)

    if settlement.has_vacant:
        vbase = ["건물주 부담(공실)", None] + ([None] if use_status else [])
        vtail = [sum(b.charges.get(n, 0) for b in settlement.vacant_bills) for n in item_names]
        vtail += [settlement.owner_borne_total] + (["-", "-"] if use_vat else [])
        if use_prev:
            vtail += ["", ""]
        ws.append(vbase + vtail)

    # 스타일: 숫자 서식, 테두리
    n_cols = len(header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=n_cols):
        for c in row:
            c.border = border
            if isinstance(c.value, (int, float)) and c.column > 2:
                c.number_format = "#,##0"
    for c in ws[total_row_idx]:
        c.font = bold
        c.fill = total_fill

    # 열 너비
    ws.column_dimensions["A"].width = 14
    for i in range(2, n_cols + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 13

    ws.freeze_panes = "C2"
    wb.save(path)
